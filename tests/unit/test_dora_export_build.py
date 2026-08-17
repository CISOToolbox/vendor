"""Perf regression (DORA H4): build_dora_xlsx still produces a valid workbook
after the CPU build was moved off the event loop.

The workbook build + zip save (~10 sheets, cell-by-cell) is pure CPU on
already-loaded data; it was wrapped in a nested sync closure run via
asyncio.to_thread so a large register can't freeze the whole vendor module.
This smoke test exercises the full async→thread→bytes path on a minimal
dataset and confirms the output is a loadable .xlsx.

Vendor has no shared db harness, so a tiny in-memory SQLite one is built here
(same JSONB/server-default shim as pilot's conftest).
"""
import io
import os
import uuid

import pytest

os.environ.setdefault("DATABASE_URL", "sqlite+aiosqlite://")
os.environ.setdefault("JWT_SECRET", "test-secret")

from sqlalchemy import JSON
from sqlalchemy.dialects.postgresql import JSONB as _JSONB
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from src.dora_export import build_dora_xlsx
from src.models import Base, Project

# ── SQLite compatibility shim (mirrors pilot/tests/conftest.py) ──
for _t in Base.metadata.tables.values():
    for _c in _t.columns:
        if _c.server_default is not None:
            _sd = str(getattr(_c.server_default, "arg", "")).lower()
            if any(k in _sd for k in ("gen_random_uuid", "now(", "::jsonb")):
                _c.server_default = None
        if isinstance(_c.type, _JSONB):
            _c.type = JSON()

_engine = create_async_engine(
    "sqlite+aiosqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool,
)
_Session = async_sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)

pytestmark = pytest.mark.asyncio


async def test_build_dora_xlsx_returns_loadable_workbook():
    async with _engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    try:
        pid = uuid.uuid4()
        async with _Session() as db:
            proj = Project(id=pid, name="MedSecure DORA")
            db.add(proj)
            await db.commit()
            blob = await build_dora_xlsx(db, pid, proj, target_currency="EUR")

        assert isinstance(blob, bytes) and len(blob) > 0
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob))
        assert len(wb.sheetnames) > 0  # at least the cover + template sheets
    finally:
        async with _engine.begin() as conn:
            await conn.run_sync(Base.metadata.drop_all)
