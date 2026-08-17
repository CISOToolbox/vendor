"""
DORA RoI XLSX export — strict EBA Data Model alignment.

Each sheet is named after the official EBA template (B_01.01 …
B_07.01) and the header row uses the official 4-digit column IDs and
labels from the EBA Data Model for DORA RoI (Nov 2024 release) and
Commission Implementing Regulation (EU) 2024/2956.

When the backing data model lacks a field required by the spec, the
column is emitted with empty values rather than skipped, so the
workbook structure remains readable by EBA-compliant validators.

Multi-currency: each arrangement keeps its own ``currency``. The
export keeps the original amounts in the canonical columns; a
non-normative sheet "Cover" surfaces the target currency for context.
A live FX feed should replace ``_FX_TO_EUR`` for production use.

Reference:
  - Commission Implementing Regulation (EU) 2024/2956 (29 Nov 2024)
  - EBA Data Model for DORA RoI (Nov 2024)
"""
from __future__ import annotations

import asyncio
import datetime as _dt
import io
import uuid
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.csv_common import xlsx_safe_workbook
from src.models import (
    DoraArrangement,
    DoraArrangementFunction,
    DoraArrangementRfe,
    DoraArrangementService,
    DoraArrangementSubcontractor,
    DoraBranch,
    DoraConsolidationScope,
    DoraEntity,
    DoraFunction,
    DoraSigner,
    DoraSubcontractor,
    Project,
    Vendor,
)


# Static FX rates (EUR base). Replace with live feed in production.
_FX_TO_EUR: dict[str, float] = {
    "EUR": 1.0, "USD": 0.92, "GBP": 1.17, "CHF": 1.04, "JPY": 0.0061,
    "CNY": 0.13, "CAD": 0.67, "AUD": 0.61, "NZD": 0.55, "SEK": 0.087,
    "NOK": 0.085, "DKK": 0.134, "PLN": 0.23, "CZK": 0.040, "HUF": 0.0026,
    "RON": 0.20, "BGN": 0.51, "HRK": 0.133, "ISK": 0.0066, "TRY": 0.027,
    "ILS": 0.25, "RUB": 0.010, "UAH": 0.022, "INR": 0.011, "SGD": 0.68,
    "HKD": 0.117, "TWD": 0.029, "KRW": 0.00067, "THB": 0.026, "IDR": 0.000058,
    "MYR": 0.20, "PHP": 0.016, "VND": 0.000037, "ZAR": 0.049, "EGP": 0.019,
    "AED": 0.25, "SAR": 0.245, "QAR": 0.252, "KWD": 3.00, "MAD": 0.092,
    "BRL": 0.16, "MXN": 0.046, "ARS": 0.001, "CLP": 0.00094, "COP": 0.00021,
    "PEN": 0.24,
}


def _convert(amount: float | None, src: str | None, target: str) -> float | None:
    if amount is None:
        return None
    src = (src or "EUR").upper()
    target = (target or "EUR").upper()
    if src == target:
        return float(amount)
    src_eur = _FX_TO_EUR.get(src)
    tgt_eur = _FX_TO_EUR.get(target)
    if src_eur is None or tgt_eur is None or tgt_eur == 0:
        return None
    return float(amount) * src_eur / tgt_eur


_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
_HEADER_FONT = Font(bold=True)


_CODE_FILL = PatternFill("solid", fgColor="EFEFEF")


def _write_sheet(wb: Workbook, name: str, cols: list[tuple[str, str]], rows: list[list[Any]]) -> None:
    """Write a sheet with the EBA two-row header convention.

    Row 1: column title (e.g. "LEI of the entity").
    Row 2: column code  (e.g. "0010").
    Row 3+: data rows.
    """
    ws = wb.create_sheet(name[:31])
    titles = [title for _cid, title in cols]
    codes = [cid for cid, _title in cols]
    ws.append(titles)
    ws.append(codes)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for cell in ws[2]:
        cell.font = Font(bold=True, italic=True, color="666666")
        cell.fill = _CODE_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append(r)
    ws.row_dimensions[1].height = 32
    for col_idx, title in enumerate(titles, start=1):
        max_len = max(
            [len(str(title)), 6]
            + [len(str(r[col_idx - 1])) if col_idx - 1 < len(r) else 0 for r in rows]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 14), 60)
    # Freeze the two-row header so users can scroll data while keeping
    # the column meaning visible.
    ws.freeze_panes = "A3"


# ── EBA Data Point Model codelists (DPM v4.0, March 2025) ─────────
# Internal codes (used in the UI and DB) are mapped to the official
# EBA codelist values expected by the Register of Information
# XBRL/CSV submission. Reference: EBA "List of possible values for
# all data fields with drop downs" (updated 3 March 2025) — DORA
# RoI ITS Reg. (EU) 2024/2956.
# Format is "eba_<TABLE>:<CODE>" with these tables in scope:
#   eba_CT  — entity type / id type (B_01.01.0040, B_01.02.0040, B_05.01.0070)
#   eba_RP  — hierarchy (B_01.02.0050)
#   eba_CO  — closed-list classifications (arrangement type, termination reason)
#   eba_BT  — Yes/No/N-A (storage, criticality, exit plan, alt TPP)
#   eba_ZZ  — graded levels (sensitivity, reliance, substitutability, reintegration, branch, impact)
#   eba_qCO — quantitative ID type (LISTIDTYPE)
#   eba_TA  — typology of activities / ICT services
#   eba_GA  — geographical area (ISO 3166-1 alpha-2 country)
#   eba_CU  — currency (ISO 4217)

# Entity types — LIST0101040 (B_01.01.0040) and LIST0102040
# (B_01.02.0040, which adds two non-financial entity codes).
_EBA_ENTITY_TYPE: dict[str, str] = {
    "credit_institution": "eba_CT:x12",
    "payment_institution": "eba_CT:x300",
    "account_information_service_provider": "eba_CT:x301",
    "electronic_money_institution": "eba_CT:x302",
    "crypto_asset_service_provider": "eba_CT:x303",
    "central_securities_depository": "eba_CT:x304",
    "trading_venue": "eba_CT:x305",
    "trade_repository": "eba_CT:x306",
    "alternative_investment_fund_manager": "eba_CT:x307",
    "data_reporting_service_provider": "eba_CT:x308",
    "insurance_undertaking": "eba_CT:x309",
    "reinsurance_undertaking": "eba_CT:x309",
    "issuer_asset_referenced_token": "eba_CT:x310",
    "ior_pension_institution": "eba_CT:x311",
    "credit_rating_agency": "eba_CT:x312",
    "benchmark_administrator": "eba_CT:x313",
    "crowdfunding_service_provider": "eba_CT:x314",
    "securitisation_repository": "eba_CT:x315",
    "other": "eba_CT:x316",
    # B_01.02-only — Non-financial counterparties
    "ict_intra_group_provider": "eba_CT:x317",
    "ict_third_party_service_provider": "eba_CT:x318",
    "subcontractor_ict_tpp": "eba_CT:x318",
    "insurance_intermediary": "eba_CT:x320",
    "ucits_management_company": "eba_CT:x639",
    "investment_firm": "eba_CT:x599",
    "central_counterparty": "eba_CT:x643",
}

# Hierarchy of entity within the group — LIST01020050 (B_01.02.0050).
_EBA_HIERARCHY: dict[str, str] = {
    "sole_entity": "eba_RP:x21",
    "parent": "eba_RP:x53",
    "ultimate_parent": "eba_RP:x53",
    "subsidiary": "eba_RP:x56",
    "intermediate_parent": "eba_RP:x551",
    "outsourcing": "eba_RP:x210",
    # No EBA code for "branch" in B_01.02.0050; branches go in B_01.03.
    "branch": "",
}

# Arrangement type — LISTB02010020 (B_02.01.0020).
_EBA_ARRANGEMENT_TYPE: dict[str, str] = {
    "standalone": "eba_CO:x1",
    "overarching": "eba_CO:x2",
    "subsequent": "eba_CO:x3",
}

# Termination reason — LISTB02020090 (B_02.02.0090). Not modelled
# yet; provided for future use.
_EBA_TERMINATION_REASON: dict[str, str] = {
    "expired_not_renewed": "eba_CO:x4",
    "breach_of_law": "eba_CO:x5",
    "impediments": "eba_CO:x6",
    "data_security_weakness": "eba_CO:x7",
    "competent_authority": "eba_CO:x8",
    "other": "eba_CO:x9",
}

# Person type — LISTB05010070 (B_05.01.0070).
_EBA_PERSON_TYPE: dict[str, str] = {
    "legal": "eba_CT:x212",
    "natural": "eba_CT:x213",
}

# Substitutability of the ICT TPP — LISTB07010050 (B_07.01.0050).
_EBA_SUBSTITUTABILITY: dict[str, str] = {
    "not_substitutable": "eba_ZZ:x959",
    "highly_complex": "eba_ZZ:x960",
    "medium_complexity": "eba_ZZ:x961",
    "easy": "eba_ZZ:x962",
}

# Reason for non-/difficult substitutability — LISTB07010060.
_EBA_SUBSTITUTABILITY_REASON: dict[str, str] = {
    "no_alternatives": "eba_ZZ:x963",
    "migration_difficulties": "eba_ZZ:x964",
    "both": "eba_ZZ:x965",
}

# Reintegration level — LISTB07010090 (B_07.01.0090).
_EBA_REINTEGRATION_LEVEL: dict[str, str] = {
    "easy": "eba_ZZ:x798",
    "difficult": "eba_ZZ:x966",
    "highly_complex": "eba_ZZ:x967",
    "not_applicable": "eba_ZZ:x0",
}

# Data sensitivity — LISTB02020170 (B_02.02.0170). EBA exposes only
# Low / Medium / High; the legacy 4-level scale is collapsed.
_EBA_DATA_SENSITIVITY: dict[str, str] = {
    "public": "eba_ZZ:x791",
    "internal": "eba_ZZ:x791",
    "low": "eba_ZZ:x791",
    "confidential": "eba_ZZ:x792",
    "medium": "eba_ZZ:x792",
    "strictly_confidential": "eba_ZZ:x793",
    "high": "eba_ZZ:x793",
}

# Level of reliance — LISTB02020180 (B_02.02.0180).
_EBA_RELIANCE_LEVEL: dict[str, str] = {
    "not_significant": "eba_ZZ:x794",
    "low": "eba_ZZ:x795",
    "material": "eba_ZZ:x796",
    "full": "eba_ZZ:x797",
}

# Impact of discontinuing — LISTB06010100 / B_07.01.0100. Same
# Low/Medium/High plus "Assessment not performed".
_EBA_IMPACT_LEVEL: dict[str, str] = {
    "low": "eba_ZZ:x791",
    "medium": "eba_ZZ:x792",
    "high": "eba_ZZ:x793",
    "not_assessed": "eba_ZZ:x799",
}

# Branch flag — LISTBB04010030 (B_04.01.0030).
_EBA_IS_BRANCH: dict[str, str] = {
    "branch": "eba_ZZ:x838",
    "not_branch": "eba_ZZ:x839",
}

# ID type for any "Type of code" column — LISTIDTYPE.
# B_02.02.0040, B_03.02.0030, B_05.01.0020, B_05.01.0040 (additional
# id), B_05.01.0120, B_05.02.0040, B_05.02.0070, B_07.01.0030.
_EBA_TYPE_OF_CODE: dict[str, str] = {
    "LEI": "eba_qCO:qx2000",
    "lei": "eba_qCO:qx2000",
    "national_id": "eba_qCO:qx2001",
    "EUID": "eba_qCO:qx2002",
    "euid": "eba_qCO:qx2002",
    "CRN": "eba_qCO:qx2003",
    "crn": "eba_qCO:qx2003",
    "VAT": "eba_qCO:qx2004",
    "vat": "eba_qCO:qx2004",
    "tax_id": "eba_qCO:qx2004",
    "passport": "eba_qCO:qx2005",
    # No "other" code in LISTIDTYPE — leave empty when unknown.
}

# ICT service typology — LISTSERVICE (B_02.02.0060, B_05.02.0020,
# B_07.01.0040). Maps internal S_01..S_21 codes to EBA S01..S19.
_EBA_ICT_SERVICE: dict[str, str] = {
    "S_01": "eba_TA:S01",
    "S_02": "eba_TA:S02",
    "S_03": "eba_TA:S03",
    "S_04": "eba_TA:S04",
    "S_05": "eba_TA:S05",
    "S_06": "eba_TA:S14",
    "S_07": "eba_TA:S07",
    "S_08": "eba_TA:S15",
    "S_09": "eba_TA:S07",
    "S_11": "eba_TA:S06",
    "S_12": "eba_TA:S16",
    "S_13": "eba_TA:S08",
    "S_14": "eba_TA:S09",
    "S_15": "eba_TA:S11",
    "S_16": "eba_TA:S10",
    "S_17": "eba_TA:S07",
    "S_18": "eba_TA:S17",
    "S_19": "eba_TA:S18",
    "S_20": "eba_TA:S19",
    # S_10 (training), S_21 (other) — no EBA equivalent.
}


def _eba(table: dict[str, str], value: Any) -> str:
    """Translate an internal code to its EBA codelist value.

    Empty / None / unknown values pass through as empty string so
    optional fields stay empty rather than being filled with a
    fallback EBA code.
    """
    if value is None or value == "":
        return ""
    s = str(value).strip()
    return table.get(s, table.get(s.lower(), ""))


# Lazy import of the JSON-loaded codelists so we can resolve eba_code
# values for large lists (e.g. licenced_activity, 131 entries).
def _eba_from_codelist(key: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        from .dora_validation import _CODELISTS  # type: ignore
    except Exception:
        return ""
    items = _CODELISTS.get(key, [])
    s = str(value)
    for it in items:
        if isinstance(it, dict) and str(it.get("code")) == s:
            return it.get("eba_code", "") or ""
    return ""


def _eba_yesno(v: Any) -> str:
    """EBA Y/N from LIST*0140 / *0080: eba_BT:x28 / eba_BT:x29."""
    if v is None or v == "":
        return ""
    return "eba_BT:x28" if v else "eba_BT:x29"


def _eba_yesno_or_na(v: Any) -> str:
    """EBA Y/N/Not-assessed from LISTB06010050 / B07010110.

    None / empty → "Assessment not performed" (eba_BT:x21).
    """
    if v is None or v == "":
        return "eba_BT:x21"
    return "eba_BT:x28" if v else "eba_BT:x29"


def _eba_country(iso2: Any) -> str:
    """Wrap an ISO-3166-1 alpha-2 country code as eba_GA:<ISO2>."""
    if iso2 is None or iso2 == "":
        return ""
    s = str(iso2).strip().upper()
    return f"eba_GA:{s}" if len(s) == 2 else ""


def _eba_currency(iso3: Any) -> str:
    """Wrap an ISO-4217 currency code as eba_CU:<ISO3>."""
    if iso3 is None or iso3 == "":
        return ""
    s = str(iso3).strip().upper()
    return f"eba_CU:{s}" if len(s) == 3 else ""


def _eba_type_of_code_for_lei(lei: str | None, fallback_type: str | None = None) -> str:
    """Helper for the very common (lei? "LEI" : additional_type) pattern."""
    if lei:
        return "eba_qCO:qx2000"
    if fallback_type:
        return _eba(_EBA_TYPE_OF_CODE, fallback_type)
    return ""


# ── Official column specs (id, title) per EBA Data Model ──────────
# Headers in the workbook are rendered as "{id} - {title}" exactly
# matching the EBA Data Model PDF.

_COLS_B0101 = [
    ("0010", "LEI of the entity maintaining the register of information"),
    ("0020", "Name of the entity"),
    ("0030", "Country of the entity"),
    ("0040", "Type of entity"),
    ("0050", "Competent Authority"),
    ("0060", "Date of the reporting"),
]

_COLS_B0102 = [
    ("0010", "LEI of the entity"),
    ("0020", "Name of the entity"),
    ("0030", "Country of the entity"),
    ("0040", "Type of entity"),
    ("0050", "Hierarchy of the entity within the group (where applicable)"),
    ("0060", "LEI of the direct parent undertaking of the financial entity"),
    ("0070", "Date of last update"),
    ("0080", "Date of integration in the Register of information"),
    ("0090", "Date of deletion in the Register of information"),
    ("0100", "Currency"),
    ("0110", "Value of total assets - of the financial entity"),
]

_COLS_B0103 = [
    ("0010", "Identification code of the branch"),
    ("0020", "LEI of the financial entity head office of the branch"),
    ("0030", "Name of the branch"),
    ("0040", "Country of the branch"),
]

_COLS_B0201 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "Type of contractual arrangement"),
    ("0030", "Overarching contractual arrangement reference number"),
    ("0040", "Currency of the amount reported"),
    ("0050", "Annual expense or estimated cost of the contractual arrangement for the past year"),
]

_COLS_B0202 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "LEI of the financial entity making use of the ICT service"),
    ("0030", "Identification code of the third-party service provider"),
    ("0040", "Type of code to identify the third-party service provider"),
    ("0050", "Function identifier"),
    ("0060", "Type of ICT services"),
    ("0070", "Start date of the contractual arrangement"),
    ("0080", "End date of the contractual arrangement"),
    ("0090", "Reason of the termination or ending of the contractual arrangement"),
    ("0100", "Notice period for the financial entity"),
    ("0110", "Notice period for the ICT third-party service provider"),
    ("0120", "Country of the governing law of the contractual arrangement"),
    ("0130", "Country of provision of the ICT services"),
    ("0140", "Storage of data"),
    ("0150", "Location of the data at rest (storage)"),
    ("0160", "Location of management of the data (processing)"),
    ("0170", "Sensitiveness of the data stored by the ICT third-party service provider"),
    ("0180", "Level of reliance on the ICT service supporting the critical or important function"),
]

_COLS_B0203 = [
    ("0010", "Contractual arrangement with ICT intra-group service provider"),
    ("0020", "Linked contractual arrangement with ICT third-party service provider"),
]

_COLS_B0301 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "LEI of the entity signing the contractual arrangement"),
]

_COLS_B0302 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "Identification code of the third-party service provider"),
    ("0030", "Type of code of the third-party service provider"),
]

_COLS_B0303 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "LEI of the intra-group entity providing ICT service"),
]

_COLS_B0401 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "LEI of the financial entity"),
    ("0030", "Is the entity making use of the ICT services a branch of a financial entity?"),
    ("0040", "Identification code of the branch"),
]

_COLS_B0501 = [
    ("0010", "Identification code of the third-party service provider"),
    ("0020", "Type of code of the third-party service provider"),
    ("0030", "Additional identification code of the third-party service provider"),
    ("0040", "Type of additional identification code of the third-party service provider"),
    ("0050", "Legal name of the third-party service provider"),
    ("0060", "Name of the ICT third-party service provider in Latin alphabet"),
    ("0070", "Type of person of the third-party service provider"),
    ("0080", "Country of the third-party service provider's headquarters"),
    ("0090", "Currency of the amount reported"),
    ("0100", "Total annual expense or estimated cost of the third-party service provider"),
    ("0110", "Identification code of the third-party service provider's ultimate parent undertaking"),
    ("0120", "Type of code of the third-party service provider's ultimate parent undertaking"),
]

_COLS_B0502 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "Type of ICT services"),
    ("0030", "Identification code of the third-party service provider"),
    ("0040", "Type of code of the third-party service provider"),
    ("0050", "Rank"),
    ("0060", "Identification code of the recipient of sub-contracted ICT services"),
    ("0070", "Type of code of the recipient of sub-contracted ICT services"),
]

_COLS_B0601 = [
    ("0010", "Function identifier"),
    ("0020", "Licenced activity"),
    ("0030", "Function name"),
    ("0040", "LEI of the financial entity"),
    ("0050", "Criticality or importance assessment"),
    ("0060", "Reasons for criticality or importance"),
    ("0070", "Date of the last assessment of criticality or importance"),
    ("0080", "Recovery time objective of the function"),
    ("0090", "Recovery point objective of the function"),
    ("0100", "Impact of discontinuing the function"),
]

_COLS_B0701 = [
    ("0010", "Contractual arrangement reference number"),
    ("0020", "Identification code of the third-party service provider"),
    ("0030", "Type of code of the third-party service provider"),
    ("0040", "Type of ICT services"),
    ("0050", "Substitutability of the ICT third-party service provider"),
    ("0060", "Reason if the ICT third-party service provider is considered not substitutable or difficult to be substitutable"),
    ("0070", "Date of the last audit on the ICT third-party service provider"),
    ("0080", "Existence of an exit plan"),
    ("0090", "Possibility of reintegration of the contracted ICT service"),
    ("0100", "Impact of discontinuing the ICT services"),
    ("0110", "Are there alternative ICT third-party service providers identified?"),
    ("0120", "Identification of alternative ICT TPP"),
]


# Templates and their human-readable titles, used for the Cover sheet.
_TEMPLATE_TITLES: list[tuple[str, str]] = [
    ("B_01.01", "Financial entity maintaining the register of information"),
    ("B_01.02", "List of financial entities within the scope of the register of information"),
    ("B_01.03", "List of branches"),
    ("B_02.01", "Contractual arrangements — General information"),
    ("B_02.02", "Contractual arrangements — Specific information"),
    ("B_02.03", "List of intra-group contractual arrangements"),
    ("B_03.01", "Entities signing the contractual arrangements (FE side)"),
    ("B_03.02", "Third-party service providers signing the contractual arrangements"),
    ("B_03.03", "Entities signing the contractual arrangements for providing ICT services"),
    ("B_04.01", "Financial entities making use of the ICT services"),
    ("B_05.01", "ICT third-party service provider"),
    ("B_05.02", "ICT service supply chains"),
    ("B_06.01", "Functions identification"),
    ("B_07.01", "Assessment of the ICT services"),
    ("B_99.01", "Entity definitions (non-normative consolidated registry)"),
]


async def build_dora_xlsx(
    db: AsyncSession,
    project_id: uuid.UUID,
    project: Project,
    *,
    target_currency: str = "EUR",
) -> bytes:
    """Build the full RoI workbook and return its bytes."""
    target_currency = (target_currency or "EUR").upper()

    entities = (await db.execute(
        select(DoraEntity).where(DoraEntity.project_id == project_id).order_by(DoraEntity.sort_order)
    )).scalars().all()
    functions = (await db.execute(
        select(DoraFunction).where(DoraFunction.project_id == project_id).order_by(DoraFunction.sort_order)
    )).scalars().all()
    branches = (await db.execute(
        select(DoraBranch).where(DoraBranch.project_id == project_id).order_by(DoraBranch.sort_order)
    )).scalars().all()
    consolidation = (await db.execute(
        select(DoraConsolidationScope).where(DoraConsolidationScope.project_id == project_id).order_by(DoraConsolidationScope.sort_order)
    )).scalars().all()
    arrangements = (await db.execute(
        select(DoraArrangement).where(DoraArrangement.project_id == project_id).order_by(DoraArrangement.sort_order)
    )).scalars().all()
    arr_rfes = (await db.execute(
        select(DoraArrangementRfe).where(DoraArrangementRfe.project_id == project_id)
    )).scalars().all()
    rfe_by_arr: dict[str, list[str]] = {}
    for link in arr_rfes:
        rfe_by_arr.setdefault(link.arrangement_id, []).append(link.rfe_id)
    arr_fns = (await db.execute(
        select(DoraArrangementFunction).where(DoraArrangementFunction.project_id == project_id)
    )).scalars().all()
    fn_by_arr: dict[str, list[str]] = {}
    for link in arr_fns:
        fn_by_arr.setdefault(link.arrangement_id, []).append(link.function_id)
    arr_svcs = (await db.execute(
        select(DoraArrangementService).where(DoraArrangementService.project_id == project_id)
    )).scalars().all()
    svc_by_arr: dict[str, list[str]] = {}
    for link in arr_svcs:
        svc_by_arr.setdefault(link.arrangement_id, []).append(link.service_code)
    signers = (await db.execute(
        select(DoraSigner).where(DoraSigner.project_id == project_id).order_by(DoraSigner.sort_order)
    )).scalars().all()
    subcontractors = (await db.execute(
        select(DoraSubcontractor).where(DoraSubcontractor.project_id == project_id).order_by(DoraSubcontractor.sort_order)
    )).scalars().all()
    sub_by_id = {s.id: s for s in subcontractors}
    subcontractor_links = (await db.execute(
        select(DoraArrangementSubcontractor)
        .where(DoraArrangementSubcontractor.project_id == project_id)
        .order_by(
            DoraArrangementSubcontractor.arrangement_id,
            DoraArrangementSubcontractor.tier,
            DoraArrangementSubcontractor.sort_order,
        )
    )).scalars().all()
    vendors = (await db.execute(
        select(Vendor).where(Vendor.project_id == project_id).order_by(Vendor.name)
    )).scalars().all()
    vendor_by_id = {v.id: v for v in vendors}
    vendor_by_lei = {(v.lei or "").upper(): v for v in vendors if v.lei}

    fn_by_id = {f.id: f for f in functions}
    rfe_by_id = {e.id: e for e in entities}
    branch_by_id = {b.id: b for b in branches}

    used_vendor_ids = {a.vendor_id for a in arrangements}
    arr_by_id = {a.id: a for a in arrangements}

    # Cumulative annual expense per vendor (B_05.01 0100), in vendor's
    # arrangement currency — note: spec asks one currency at vendor
    # level (0090). We keep the most-frequent currency among the
    # vendor's arrangements.
    vendor_total_by_id: dict[str, float] = {}
    vendor_cur_by_id: dict[str, str] = {}
    cur_count: dict[str, dict[str, int]] = {}
    for a in arrangements:
        if a.annual_cost_amount is None:
            continue
        vendor_total_by_id[a.vendor_id] = vendor_total_by_id.get(a.vendor_id, 0.0) + float(a.annual_cost_amount)
        cur = (a.currency or "EUR").upper()
        bucket = cur_count.setdefault(a.vendor_id, {})
        bucket[cur] = bucket.get(cur, 0) + 1
    for vid, counts in cur_count.items():
        vendor_cur_by_id[vid] = max(counts, key=counts.get)

    # The workbook build + zip save is pure CPU on already-loaded data
    # (no DB access — async SQLAlchemy would raise on a lazy load, and
    # this function works today, proving every attribute is materialized).
    # Run it off the event loop so a large register doesn't freeze the
    # whole vendor module (health checks, Pilot stats, autosaves) for the
    # duration of the export.
    def _build() -> bytes:
        wb = Workbook()
        cover = wb.active
        cover.title = "Cover"
        cover["A1"] = "DORA Register of Information"
        cover["A1"].font = Font(bold=True, size=14)
        cover["A3"] = "Project"
        cover["B3"] = project.name or ""
        cover["A4"] = "Reporting period"
        cover["B4"] = entities[0].reporting_period if entities else ""
        cover["A5"] = "Target currency (illustrative FX)"
        cover["B5"] = target_currency
        cover["A6"] = "EBA reference"
        cover["B6"] = "Reg. (EU) 2024/2956 — EBA Data Model for DORA RoI"
        cover["A8"] = "Templates"
        cover["A8"].font = Font(bold=True)
        for i, (tid, label) in enumerate(_TEMPLATE_TITLES, start=9):
            cover.cell(row=i, column=1, value=tid)
            cover.cell(row=i, column=2, value=label)
        cover.column_dimensions["A"].width = 12
        cover.column_dimensions["B"].width = 80

        today = _dt.date.today().isoformat()

        # ── B_01.01 — Entity maintaining the register (1 row) ──────────
        register_holder = next((e for e in entities if (e.hierarchy or "") == "parent"), None)
        if register_holder is None and entities:
            register_holder = entities[0]
        holder_rows: list[list[Any]] = []
        if register_holder:
            e = register_holder
            holder_rows.append([
                e.lei or "",
                e.name or "",
                _eba_country(e.country_iso2),
                _eba(_EBA_ENTITY_TYPE, e.entity_type),
                e.competent_authority or "",
                e.reporting_period or today,
            ])
        _write_sheet(wb, "B_01.01", _COLS_B0101, holder_rows)

        # ── B_01.02 — List of FEs in scope ─────────────────────────────
        b0102_rows: list[list[Any]] = []
        for e in entities:
            b0102_rows.append([
                e.lei or "",
                e.name or "",
                _eba_country(e.country_iso2),
                _eba(_EBA_ENTITY_TYPE, e.entity_type),
                _eba(_EBA_HIERARCHY, e.hierarchy),
                e.parent_lei or "",
                (e.updated_at.date().isoformat() if e.updated_at else ""),
                (e.created_at.date().isoformat() if e.created_at else ""),
                "",  # date of deletion — not modelled
                _eba_currency(e.total_assets_currency or "EUR") if e.total_assets is not None else "",
                e.total_assets if e.total_assets is not None else "",
            ])
        for c in consolidation:
            # Consolidation scope rows are not strictly RFEs; entity_type
            # and hierarchy are left empty since the data model does not
            # store an EBA entity_type for consolidation entries.
            b0102_rows.append([
                c.entity_lei or "",
                c.entity_name or "",
                _eba_country(c.country_iso2),
                "",  # entity_type — not modelled on consolidation rows
                "",  # hierarchy — not declared on consolidation rows
                "",
                "", "", "", "", "",
            ])
        _write_sheet(wb, "B_01.02", _COLS_B0102, b0102_rows)

        # ── B_01.03 — Branches ─────────────────────────────────────────
        b0103_rows: list[list[Any]] = []
        for b in branches:
            head_rfe = rfe_by_id.get(b.rfe_id or "")
            b0103_rows.append([
                b.branch_code or b.id,
                (head_rfe.lei if head_rfe else ""),
                b.name or "",
                _eba_country(b.country_iso2),
            ])
        _write_sheet(wb, "B_01.03", _COLS_B0103, b0103_rows)

        # ── B_02.01 — Contractual arrangements (general info) ──────────
        # 0030 Reference of the overarching arrangement: only populated when
        # the arrangement is itself a "subsequent" arrangement of an
        # overarching one. We resolve the parent's arrangement_reference
        # rather than emitting the internal id so the value is human-readable.
        arr_ref_by_id = {a.id: (a.arrangement_reference or a.id) for a in arrangements}
        b0201_rows: list[list[Any]] = []
        for a in arrangements:
            parent_ref = ""
            if a.parent_arrangement_id:
                parent_ref = arr_ref_by_id.get(a.parent_arrangement_id, a.parent_arrangement_id)
            b0201_rows.append([
                a.arrangement_reference or a.id,
                _eba(_EBA_ARRANGEMENT_TYPE, a.arrangement_type),
                parent_ref,  # 0030 overarching arrangement ref
                _eba_currency(a.currency),
                a.annual_cost_amount if a.annual_cost_amount is not None else "",
            ])
        _write_sheet(wb, "B_02.01", _COLS_B0201, b0201_rows)

        # ── B_02.02 — Contractual arrangements (specific info) ─────────
        # PK: (ref, FE LEI, TPSP id, function id) → emit cartesian product.
        b0202_rows: list[list[Any]] = []
        for a in arrangements:
            v = vendor_by_id.get(a.vendor_id)
            rfes = rfe_by_arr.get(a.id, []) or [None]
            fns = fn_by_arr.get(a.id, []) or [None]
            # ITS B_02.02.0060 takes one service code per row. An arrangement
            # may declare multiple service codes (junction). Fall back to the
            # legacy single-string nature_of_service when the junction is
            # empty so existing data still exports.
            svc_codes = svc_by_arr.get(a.id, [])
            if not svc_codes:
                svc_codes = [a.nature_of_service] if a.nature_of_service else [None]
            for rfe_id in rfes:
                for fn_id in fns:
                    for svc in svc_codes:
                        rfe = rfe_by_id.get(rfe_id) if rfe_id else None
                        # B.06.01.0010 — prefer the user-editable function code
                        # (free identifier) and fall back to the internal id.
                        fn_code = ""
                        if fn_id:
                            fn_obj = fn_by_id.get(fn_id)
                            fn_code = (getattr(fn_obj, "code", "") or fn_id) if fn_obj else fn_id
                        b0202_rows.append([
                            a.arrangement_reference or a.id,
                            (rfe.lei if rfe else ""),
                            (v.lei if v and v.lei else (v.id if v else "")),
                            _eba_type_of_code_for_lei(v.lei if v else None),
                            fn_code,
                            _eba(_EBA_ICT_SERVICE, svc) if svc else "",
                            a.start_date or "",
                            a.end_date or "",
                            _eba(_EBA_TERMINATION_REASON, a.termination_reason),
                            a.notice_period_days if a.notice_period_days is not None else "",
                            a.notice_period_tpsp_days if a.notice_period_tpsp_days is not None else "",
                            _eba_country(a.governing_law_country),
                            _eba_country(a.jurisdiction_country),
                            _eba_yesno(bool(a.data_storage_country)),
                            _eba_country(a.data_storage_country),
                            _eba_country(a.data_processing_country),
                            _eba(_EBA_DATA_SENSITIVITY, a.data_sensitivity),
                            _eba(_EBA_RELIANCE_LEVEL, a.reliance_level),
                        ])
        _write_sheet(wb, "B_02.02", _COLS_B0202, b0202_rows)

        # ── B_02.03 — Intra-group arrangements ─────────────────────────
        rfe_lei_set = {(e.lei or "").upper() for e in entities if e.lei}
        b0203_rows: list[list[Any]] = []
        for a in arrangements:
            v = vendor_by_id.get(a.vendor_id)
            if not v:
                continue
            is_intra = False
            if v.ultimate_parent_id:
                pv = vendor_by_id.get(v.ultimate_parent_id)
                if pv and pv.lei and pv.lei.upper() in rfe_lei_set:
                    is_intra = True
            if not is_intra:
                continue
            b0203_rows.append([
                a.arrangement_reference or a.id,
                "",  # linked external arrangement — not modelled
            ])
        _write_sheet(wb, "B_02.03", _COLS_B0203, b0203_rows)

        # ── B_03.01 — Entities signing (FE side) ───────────────────────
        b0301_rows: list[list[Any]] = []
        for s in signers:
            if (s.signer_role or "").lower() == "tpp":
                continue
            a = arr_by_id.get(s.arrangement_id)
            b0301_rows.append([
                (a.arrangement_reference if a else s.arrangement_id),
                s.signer_lei or "",
            ])
        _write_sheet(wb, "B_03.01", _COLS_B0301, b0301_rows)

        # ── B_03.02 — TPSP signers ─────────────────────────────────────
        b0302_rows: list[list[Any]] = []
        for a in arrangements:
            v = vendor_by_id.get(a.vendor_id)
            if v:
                b0302_rows.append([
                    a.arrangement_reference or a.id,
                    v.lei or v.id or "",
                    _eba_type_of_code_for_lei(v.lei, v.additional_id_type),
                ])
        for s in signers:
            if (s.signer_role or "").lower() != "tpp":
                continue
            a = arr_by_id.get(s.arrangement_id)
            b0302_rows.append([
                (a.arrangement_reference if a else s.arrangement_id),
                s.signer_lei or "",
                _eba_type_of_code_for_lei(s.signer_lei),
            ])
        _write_sheet(wb, "B_03.02", _COLS_B0302, b0302_rows)

        # ── B_03.03 — Intra-group entity providing ICT (placeholder) ───
        _write_sheet(wb, "B_03.03", _COLS_B0303, [])

        # ── B_04.01 — Entities making use of ICT services ──────────────
        b0401_rows: list[list[Any]] = []
        for link in arr_rfes:
            a = arr_by_id.get(link.arrangement_id)
            rfe = rfe_by_id.get(link.rfe_id)
            b0401_rows.append([
                (a.arrangement_reference if a else link.arrangement_id),
                (rfe.lei if rfe else ""),
                _EBA_IS_BRANCH["not_branch"],  # RFE rows are FEs, not branches
                "",
            ])
        # Branch usage rows — emit when a branch is explicitly recorded as
        # a user. The current model does not link branches to arrangements
        # directly; left empty so the sheet stays correct rather than wrong.
        _write_sheet(wb, "B_04.01", _COLS_B0401, b0401_rows)

        # ── B_05.01 — TPSP catalog ─────────────────────────────────────
        b0501_rows: list[list[Any]] = []
        for v in vendors:
            if v.id not in used_vendor_ids:
                continue
            ult = vendor_by_id.get(v.ultimate_parent_id) if v.ultimate_parent_id else None
            b0501_rows.append([
                v.lei or v.id or "",
                _eba_type_of_code_for_lei(v.lei),
                v.additional_id_value or "",
                _eba(_EBA_TYPE_OF_CODE, v.additional_id_type),
                v.name or "",
                v.legal_name_latin or "",
                _eba(_EBA_PERSON_TYPE, v.person_type),
                _eba_country(v.country_iso2),
                _eba_currency(vendor_cur_by_id.get(v.id, "")),
                vendor_total_by_id.get(v.id, ""),
                (ult.lei if ult and ult.lei else (v.ultimate_parent_id or "")),
                _eba_type_of_code_for_lei(ult.lei if ult else None),
            ])
        _write_sheet(wb, "B_05.01", _COLS_B0501, b0501_rows)

        # ── B_05.02 — Supply chains ────────────────────────────────────
        # EBA DORA RoI FAQ (28 March 2025), Q66 / Q69 / Q81 / Q98:
        #   • rank = 1 → row describes the *direct* ICT TPP. Column 0060 is
        #     a primary key and cannot be empty, so 0060 must repeat the
        #     value of 0030 (the TPSP LEI).
        #   • rank ≥ 2 → row describes a subcontractor. 0030 = subcontractor
        #     LEI (provider at this rank), 0060 = LEI of the entity at rank
        #     n‑1 (the direct TPSP at tier=2, the parent subcontractor for
        #     tier ≥ 3).
        #   • All direct TPSPs must appear (Q69.a/b), so emit a rank=1 row
        #     per (arrangement × service) regardless of whether the service
        #     supports a critical/important function.
        b0502_rows: list[list[Any]] = []

        # Rank=1 rows — one per (arrangement × service), mirroring the B_02.02
        # cartesian expansion so service codes stay consistent across sheets.
        for a in arrangements:
            v = vendor_by_id.get(a.vendor_id)
            if not v:
                continue
            vendor_lei = (v.lei if v.lei else v.id) or ""
            vendor_code_type = _eba_type_of_code_for_lei(v.lei)
            svc_codes = svc_by_arr.get(a.id, [])
            if not svc_codes and a.nature_of_service:
                svc_codes = [a.nature_of_service]
            for svc in svc_codes:
                b0502_rows.append([
                    a.arrangement_reference or a.id,
                    _eba(_EBA_ICT_SERVICE, svc) if svc else "",
                    vendor_lei,
                    vendor_code_type,
                    1,
                    vendor_lei,            # 0060 = 0030 at rank 1 (EBA Q66)
                    vendor_code_type,
                ])

        # Rank ≥ 2 rows — one per subcontractor_link. Recipient is the entity
        # at rank n‑1: vendor when tier=2, parent_subcontractor otherwise.
        for link in subcontractor_links:
            a = arr_by_id.get(link.arrangement_id)
            v = vendor_by_id.get(a.vendor_id) if a else None
            sub = sub_by_id.get(link.subcontractor_id)
            if not sub:
                continue
            sub_lei = (sub.lei if sub.lei else sub.id) or link.subcontractor_id
            sub_code_type = _eba_type_of_code_for_lei(sub.lei)
            tier = link.tier if link.tier is not None else 2
            if tier <= 2 or not link.parent_subcontractor_id:
                recipient_lei = (v.lei if v and v.lei else (v.id if v else "")) or ""
                recipient_code_type = _eba_type_of_code_for_lei(v.lei if v else None)
            else:
                parent_sub = sub_by_id.get(link.parent_subcontractor_id)
                if parent_sub:
                    recipient_lei = (parent_sub.lei if parent_sub.lei else parent_sub.id) or link.parent_subcontractor_id
                    recipient_code_type = _eba_type_of_code_for_lei(parent_sub.lei)
                else:
                    recipient_lei = link.parent_subcontractor_id or ""
                    recipient_code_type = _eba_type_of_code_for_lei(None)
            b0502_rows.append([
                (a.arrangement_reference if a else link.arrangement_id),
                _eba(_EBA_ICT_SERVICE, link.service_provided),
                sub_lei,             # 0030 = provider at rank n (subcontractor)
                sub_code_type,
                tier,
                recipient_lei,       # 0060 = entity at rank n-1 (EBA Q81)
                recipient_code_type,
            ])
        _write_sheet(wb, "B_05.02", _COLS_B0502, b0502_rows)

        # ── B_06.01 — Functions identification ─────────────────────────
        # Functions are project-level in this model, not RFE-specific.
        # The "LEI of the financial entity" column is filled with the
        # register holder's LEI as a sensible default.
        holder_lei = (register_holder.lei if register_holder else "") or ""
        b0601_rows: list[list[Any]] = []
        for f in functions:
            b0601_rows.append([
                (getattr(f, "code", "") or "") or f.id,
                _eba_from_codelist("licenced_activity", f.business_line),
                f.name or "",
                holder_lei,
                _eba_yesno_or_na(f.is_critical_or_important),
                f.criticality_rationale or "",
                f.last_assessment_date or "",  # B_06.01.0070
                f.recovery_time_objective_h if f.recovery_time_objective_h is not None else "",
                f.recovery_point_objective_h if f.recovery_point_objective_h is not None else "",
                _eba(_EBA_IMPACT_LEVEL, f.impact_tolerance_description),
            ])
        _write_sheet(wb, "B_06.01", _COLS_B0601, b0601_rows)

        # ── B_07.01 — Assessment of the ICT services ───────────────────
        # PK is (arrangement_reference × ICT service). One row per declared
        # service so each service inherits the arrangement's substitutability /
        # reintegration / exit assessment. Falls back to legacy nature_of_service
        # when no junction row exists.
        b0701_rows: list[list[Any]] = []
        for a in arrangements:
            v = vendor_by_id.get(a.vendor_id)
            svc_codes = svc_by_arr.get(a.id, [])
            if not svc_codes:
                svc_codes = [a.nature_of_service] if a.nature_of_service else [None]
            # 0110 alternative TPP "yes/no/na" boolean is derived from whether
            # an alternative TPSP id was recorded; the actual identifier flows
            # into the column when present, otherwise leave empty.
            alt_present = bool((a.alternative_tpp_id or "").strip())
            for svc in svc_codes:
                b0701_rows.append([
                    a.arrangement_reference or a.id,
                    (v.lei if v and v.lei else (v.id if v else "")),
                    _eba_type_of_code_for_lei(v.lei if v else None),
                    _eba(_EBA_ICT_SERVICE, svc) if svc else "",
                    _eba(_EBA_SUBSTITUTABILITY, a.substitutability_level),
                    _eba(_EBA_SUBSTITUTABILITY_REASON, a.substitutability_reason),
                    a.last_audit_date or "",
                    _eba_yesno(bool(a.exit_strategy_documented)),
                    _eba(_EBA_REINTEGRATION_LEVEL, a.reintegration_level),
                    _eba(_EBA_IMPACT_LEVEL, a.impact_discontinuing_level),  # B_07.01.0100
                    _eba_yesno_or_na(alt_present),                          # B_07.01.0110 yes/no/na
                    a.alternative_tpp_id or "",                             # B_07.01 alternative TPP ref
                ])
        _write_sheet(wb, "B_07.01", _COLS_B0701, b0701_rows)

        # ── B_99.01 — Non-normative consolidated registry ──────────────
        # Useful internal cross-reference of every legal entity mentioned.
        b9901_rows: list[list[Any]] = []
        for e in entities:
            b9901_rows.append([e.id, "RFE", e.lei or "", e.name or "", e.country_iso2 or "", e.parent_lei or ""])
        for b in branches:
            b9901_rows.append([b.id, "Branch", b.lei or "", b.name or "", b.country_iso2 or "", ""])
        for c in consolidation:
            b9901_rows.append([c.id, "ConsolidationScope", c.entity_lei or "", c.entity_name or "", c.country_iso2 or "", ""])
        for v in vendors:
            b9901_rows.append([v.id, "TPSP", v.lei or "", v.name or "", v.country_iso2 or "", v.ultimate_parent_id or ""])
        for s in subcontractors:
            b9901_rows.append([s.id, "Subcontractor", s.lei or "", s.name or "", s.country_iso2 or "", ""])
        _write_sheet(
            wb,
            "B_99.01",
            [
                ("0010", "id"),
                ("0020", "kind"),
                ("0030", "lei"),
                ("0040", "name"),
                ("0050", "country_iso2"),
                ("0060", "parent_ref"),
            ],
            b9901_rows,
        )

        # Mark unused references explicit (silences static analysis):
        _ = (vendor_by_lei, branch_by_id, fn_by_id)

        # Formula-injection guard. Every value above comes from user-editable
        # vendor/entity records, and openpyxl turns any string starting with
        # "=" into a real <f> formula that Excel evaluates on open. Neutralise
        # the whole workbook in one pass, right before writing it, so a sheet
        # added later is covered without touching this line again. Values are
        # preserved byte for byte — see src/csv_common.xlsx_safe_workbook.
        xlsx_safe_workbook(wb)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    return await asyncio.to_thread(_build)
